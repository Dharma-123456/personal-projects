from bs4 import BeautifulSoup
import requests
import openpyxl

excel=openpyxl.Workbook()
print(excel.sheetnames)
sheet =excel.active
sheet.title= 'Top rated Movies'
print(excel.sheetnames)
sheet.append(["Movie rank","Movie name","Year of release","IMDB rank"])



try:
    source=requests.get('https://www.imdb.com/chart/top/')
    source.raise_for_status()
    soup= BeautifulSoup(source.text,'html.parser')
    # print(soup)

    movies=soup.find('tbody',class_="lister-list").find_all('tr')
    # print(len(movies))

    for movie in movies:
        name = movie.find('td', class_="titleColumn").a.text

        rank= movie.find('td',class_="titleColumn").get_text(strip=True).split('.')[0]

        year= movie.find('td', class_="titleColumn").span.text.strip('()')

        rating =movie.find('td',class_='ratingColumn imdbRating').strong.text

        print(name ,rank ,year ,rating)
        sheet.append([rank,name,year,rating])



except Exception as e:
    print(e) 

excel.save('IMDB movie ratings.xlsx')